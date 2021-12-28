import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

book = load_workbook('titanic.xlsx')
df = pd.read_excel('titanic.xlsx')
col_names = df.columns
column_index = [col_names.get_loc(col_name) for col_name in col_names]
# combination list of column names & column index
col_indx = tuple(zip(col_names,column_index))

def conditional_formatting(sheetname,hexacode):
    for i in range(0, len(col_indx)):
        col_header = col_indx[i][0]
        col_index = col_indx[i][1]
        col_numric = df[col_header].dtype.kind in 'biufc'
        rows = df[col_header].index.to_list()
        if col_index >= 2:
            if col_numric == True:
                for row in rows:
                    book[sheetname].cell(row=row + 2, column=col_index+1).fill = PatternFill(fgColor=hexacode,fill_type='solid')
        book.save('titanic.xlsx')
conditional_formatting("data","0000FF")